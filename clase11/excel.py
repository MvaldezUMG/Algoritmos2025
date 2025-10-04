import openpyxl

def crear_libro(nombre):
    wb = openpyxl.Workbook()
    #Usamos la hoja de trabajo activa
    worksheet = wb.active
    worksheet["A1"] = "Carnet"
    worksheet["B1"] = "Nombre"
    worksheet["C1"] = "Carrera"

    #Guardamos la hoja
    wb.save(nombre)

def leer_datos(nombre):
    #Abrimos el archivo desde el disco
    libro = openpyxl.load_workbook(nombre)
    nombre_primera_hoja = libro.sheetnames[0]
    hoja = libro[nombre_primera_hoja]
    for numero_fila in range(2, hoja.max_row + 1):
        print("Carnet: " + hoja["A" + str(numero_fila)].value)
        print("Nombre: " + hoja["B" + str(numero_fila)].value)
        print("Carrera: " + hoja["C" + str(numero_fila)].value)

def escribir_datos(nombre_archivo, carnet, nombre, carrera):
    libro = openpyxl.load_workbook(nombre_archivo)
    nombre_primera_hoja = libro.sheetnames[0]
    hoja = libro[nombre_primera_hoja]
    proxima_fila = hoja.max_row + 1
    hoja["A" + str(proxima_fila)] = carnet
    hoja["B" + str(proxima_fila)] = nombre
    hoja["C" + str(proxima_fila)] = carrera
    #Guardamos en disco
    libro.save(nombre_archivo)

def editar_datos(nombre_archivo, carnet, nombre, carrera):
    libro = openpyxl.load_workbook(nombre_archivo)
    nombre_primera_hoja = libro.sheetnames[0]
    fila_encontrada = -1
    hoja = libro[nombre_primera_hoja]
    for fila in range(2, hoja.max_row + 1):
        if hoja["A" + str(fila)].value == carnet:
            fila_encontrada = fila
            break
    if fila_encontrada > -1:
        hoja["B"+ str(fila_encontrada)]= nombre
        hoja["C" + str(fila_encontrada)] = carrera
        libro.save(nombre_archivo)

while True:
    nombre_archivo = ""
    print("1. Crear libro excel")
    print("2. Leer datos de libro excel")
    print("3. Agregar datos a libro excel")
    print("4. Editar datos en libro excel")
    print("0. Salir")

    opcion = input("Ingrese la opcion: \n")
    if opcion == "1":
        nombre_archivo = input("Ingrese el nombre del archivo\n")
        crear_libro(nombre + ".xlsx")
    if opcion == "2":
        leer_datos(nombre_archivo + ".xlsx")
    if opcion == "3":
        carnet = input("ingrese el carnet del estudiante\n")
        nombre = input("ingrese el nombre del estudiante\n")
        carrera = input("ingrese la carrera del estudiante \n")
        escribir_datos(nombre_archivo + ".xlsx", carnet, nombre, carrera)
    if opcion == "4":
        carnet = input("ingrese el carnet del estudiante\n")
        nombre = input("ingrese el nombre del estudiante\n")
        carrera = input("ingrese la carrera del estudiante \n")
        editar_datos(nombre_archivo + ".xlsx", carnet, nombre, carrera)        
    if opcion == "0":
        break