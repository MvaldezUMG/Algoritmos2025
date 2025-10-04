import openpyxl 

def leer_alumnos():
    #creamos un libro
    libro = openpyxl.load_workbook("estudiantes.xlsx")
    #abrir una hoja
    hoja = libro["alumnos"]
    alumnos = []
    for numero_fila in range(2, hoja.max_row + 1):
        alumno = {
            "nombre": hoja["A"+ str(numero_fila)].value,
            "edad": hoja["B"+str(numero_fila)].value
        }
        alumnos.append(alumno)
    return alumnos

#datos = leer_alumnos()
#print(datos)

def guardar_alumno():
    #creamos un libro
    libro = openpyxl.load_workbook("estudiantes.xlsx")
    #abrir una hoja
    hoja = libro["alumnos"]
    nombre = input("ingrese el nombre \n")
    edad = input("ingrese la edad\n")
    proxima_fila = hoja.max_row + 1
    hoja["A"+str(proxima_fila)].value = nombre
    hoja["B"+str(proxima_fila)].value = edad
    #Guardamos en disco
    libro.save("estudiantes.xlsx")

#escribir_datos()

while True:
    print("1. Leer alumnos")
    print("2. Agregar alumno")
    print("3. Salir")
    print("Ingrese una opcion:")
    opcion = input()

    if opcion == "1":
        alumnos = leer_alumnos()
        for alumno in alumnos:
            print("nombre:" + alumno["nombre"] + ", edad: " + str(alumno["edad"]))
    if opcion == "2":
        guardar_alumno()
    if opcion == "3":
        break